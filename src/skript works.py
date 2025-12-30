import os
import re
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from config_paths import OUTPUT_FRANVARO_DIR

# === Sökvägar ===
INPUT_PATH = OUTPUT_FRANVARO_DIR / "franvaro.xls"
OUTPUT_DIR = OUTPUT_FRANVARO_DIR


def safe_sheet_name(name: str) -> str:
    """Excel-bladnamn: max 31 tecken + förbjudna tecken."""
    name = re.sub(r"[:\/\?\*\[\]]", "_", str(name))
    return name[:31]


def convert_percent(col: pd.Series) -> pd.Series:
    return pd.to_numeric(
        col.astype(str)
           .str.replace("%", "", regex=False)
           .str.replace(",", ".", regex=False)
           .str.replace("\xa0", "", regex=False)
           .str.extract(r"(\d+\.?\d*)")[0],
        errors="coerce"
    )


def extrahera_arskurs(klass: str) -> str:
    if isinstance(klass, str):
        klass = klass.strip()
        if klass.lower().startswith("agsä"):
            return klass
        match = re.search(r"\d", klass)
        if match:
            return f"Åk {match.group()}"
    return "Åk F"


def bygg_summering(df_in: pd.DataFrame) -> pd.DataFrame:
    """Samma översikt som tidigare, men för valfritt urval (kommun/skola)."""

    def get_total_kategori(narvaro_pct):
        franvaro = 100 - narvaro_pct if pd.notna(narvaro_pct) else None
        if franvaro is None:
            return None
        if franvaro > 50.0: return "50,1--%"
        elif franvaro > 30.0: return "30,1-50,0%"
        elif franvaro > 15.0: return "15,1-30,0%"
        elif franvaro > 5.0: return "5,1-15,0%"
        else: return "0,0-5,0%"

    def get_ogiltig_kategori(p):
        if pd.isna(p): return None
        if p > 15.0: return "15,1--%"
        elif p > 5.0: return "5,1-15,0%"
        elif p >= 1.0: return "1,0-5,0%"
        else: return None

    arskurser = sorted(df_in["årskurs"].dropna().unique())
    summering = pd.DataFrame(index=arskurser)

    totalkategorier = ["0,0-5,0%", "5,1-15,0%", "15,1-30,0%", "30,1-50,0%", "50,1--%"]
    ogiltigkategorier = ["1,0-5,0%", "5,1-15,0%", "15,1--%"]

    for kat in totalkategorier:
        summering[f"Total frånvaro {kat}"] = 0
    for kat in ogiltigkategorier:
        summering[f"Ogiltig frånvaro {kat}"] = 0
    summering["Elevantal"] = 0

    for _, row in df_in.iterrows():
        ak = row["årskurs"]
        total_kat = get_total_kategori(row["närvaro_pct"])
        ogiltig_kat = get_ogiltig_kategori(row["ogiltig_frånvaro_pct"])
        if total_kat:
            summering.at[ak, f"Total frånvaro {total_kat}"] += 1
        if ogiltig_kat:
            summering.at[ak, f"Ogiltig frånvaro {ogiltig_kat}"] += 1
        summering.at[ak, "Elevantal"] += 1

    return summering


def skriv_df_till_sheet(ws, df: pd.DataFrame):
    for r in dataframe_to_rows(df, index=False, header=True):
        if any(str(cell).strip() not in ["", "nan", "NaN"] for cell in r):
            ws.append(r)


def skriv_summering_till_sheet(ws, summering: pd.DataFrame):
    for r in dataframe_to_rows(summering, index=True, header=True):
        ws.append(r)


def formatera_data_sheet(ws):
    fill_colors = {'A': "FFFFFF", 'B': "C0C0C0", 'C': "C4D79B", 'D': "FFFF99", 'E': "FF9999"}
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=min(5, ws.max_column)):
        for cell in row:
            col_letter = cell.column_letter
            cell.fill = PatternFill(start_color=fill_colors.get(col_letter, "FFFFFF"),
                                    end_color=fill_colors.get(col_letter, "FFFFFF"),
                                    fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=(cell.col_idx == 1))
            cell.border = border


def autosize_columns(wb: Workbook):
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = max_len + 2


# === STEG 1: Läs in och tolka ===
# OBS: franvaro.xls har nu första kolumnen "skola"
raw = pd.read_excel(INPUT_PATH, header=None)
data = []
aktuell_klass = None

for row in raw.itertuples(index=False):
    skola = str(row[0]).strip() if len(row) > 0 else ""
    row_list = list(row)

    # Fånga "Klass:" i valfri kolumn (för att tåla varierande export)
    klass_cell = next((c for c in row_list if isinstance(c, str) and "Klass:" in c), None)
    if klass_cell:
        aktuell_klass = klass_cell.split(":", 1)[1].strip()
        continue

    if not aktuell_klass:
        continue

    rest = list(row_list[1:])  # allt utom skola

    def is_blank(v):
        if v is None:
            return True
        s = str(v).strip().lower()
        return s == "" or s == "nan"

    # Droppa alla ledande tomkolumner efter skola ("" eller NaN)
    while rest and is_blank(rest[0]):
        rest = rest[1:]

    # Filtrera bort rubrikrader (kräv minst ett numeriskt värde i mätkolumnerna)
    has_numeric = any(pd.to_numeric(val, errors="coerce") == pd.to_numeric(val, errors="coerce") for val in rest)
    if not has_numeric:
        continue

    # EXAKT 10 kolumner efter skola: namn, personnr, undv_tid, lekt, n_min, gf_min, f_min, n_pct, gf_pct, f_pct
    if len(rest) < 10:
        rest += [""] * (10 - len(rest))
    else:
        rest = rest[:10]

    data.append([skola, aktuell_klass] + rest)

if not data:
    print("⚠️ Hittade inga datarader att skriva ut. Kontrollera att 'franvaro.xls' har kolumn 1=skola, samt rader med 'Klass:' i någon kolumn, och minst en datarad med numeriskt värde.")
    sys.exit(1)

df = pd.DataFrame(data).iloc[:, :12]
df.columns = [
    "skola", "klass", "namn", "personnr", "undv_tid", "lekt",
    "n_min", "gf_min", "f_min", "n_pct", "gf_pct", "f_pct"
]

# Ta bort rader där personnr är tom eller där rubriker läckt med
df = df[
    df["personnr"].notna() &
    ~df["personnr"].astype(str).str.lower().str.contains("personnr|namn|undv_tid")
]

df["årskurs"] = df["klass"].apply(extrahera_arskurs)
df["närvaro_pct"] = convert_percent(df["n_pct"])
df["ogiltig_frånvaro_pct"] = convert_percent(df["f_pct"])

# Ta bort helt tomma mätvärden (om båda är NaN)
df = df[~(df["närvaro_pct"].isna() & df["ogiltig_frånvaro_pct"].isna())]

summering_kommun = bygg_summering(df)

# === STEG 3: Skapa Excel ===
wb = Workbook()
ws_kommun_data = wb.active
ws_kommun_data.title = "Rensad data - Kommun"
skriv_df_till_sheet(ws_kommun_data, df)
formatera_data_sheet(ws_kommun_data)

# Kommun: översikt
ws_kommun_sum = wb.create_sheet("Översikt - Kommun")
skriv_summering_till_sheet(ws_kommun_sum, summering_kommun)

# Skolvisa flikar
for skola in sorted(df["skola"].dropna().unique()):
    df_s = df[df["skola"] == skola].copy()
    summering_s = bygg_summering(df_s)

    ws_data = wb.create_sheet(safe_sheet_name(f"{skola} - Rensad data"))
    skriv_df_till_sheet(ws_data, df_s)
    formatera_data_sheet(ws_data)

    ws_sum = wb.create_sheet(safe_sheet_name(f"{skola} - Översikt"))
    skriv_summering_till_sheet(ws_sum, summering_s)

# Autosize sist (så det gäller alla blad)
autosize_columns(wb)

# === Spara ===
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "franvaro_rensad_kategoriserad.xlsx")
wb.save(OUTPUT_PATH)
print(f"✔️ Klar! Filen sparades till {OUTPUT_PATH}")

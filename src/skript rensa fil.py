import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Läs in .xls-filen
df = pd.read_excel("franvaro.xls", header=None)  # Kräver 'xlrd'

# Rensa de 7 första raderna
df = df.iloc[7:].reset_index(drop=True)

# Ta bort rader där kolumn B (index 1) börjar med "Namn" eller "Klass"
mask = df[1].astype(str).str.startswith(("Namn", "Klass"))
df = df[~mask].reset_index(drop=True)

# Ta bort kolumner D till H (index 3–7) om de finns
df.drop(df.columns[3:8], axis=1, inplace=True, errors='ignore')

# Ta bort kolumn A (index 0) om den finns kvar
df.drop(df.columns[0], axis=1, inplace=True, errors='ignore')

# Dela på tabbtecken i alla kolumner som innehåller tabbar
for col in df.columns:
    if df[col].astype(str).str.contains('\t').any():
        expanded = df[col].astype(str).str.split('\t', expand=True)
        df = pd.concat([df.drop(columns=[col]), expanded], axis=1)

# Skapa ny arbetsbok
wb = Workbook()
ws = wb.active

# Skriv in DataFrame till kalkylblad
for r in dataframe_to_rows(df, index=False, header=False):
    ws.append(r)

# Formatera kolumner A till E (eller färre om färre finns)
fill_colors = {
    'A': "FFFFFF",  # Ingen färg
    'B': "C0C0C0",  # Ljusgrå
    'C': "C4D79B",  # Grönaktig
    'D': "FFFF99",  # Gul
    'E': "FF9999",  # Rödaktig
}

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=min(5, ws.max_column)):
    for cell in row:
        col_letter = cell.column_letter
        fill_color = fill_colors.get(col_letter, "FFFFFF")
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=(cell.col_idx == 1))  # Första kolumnen fetstil
        cell.border = border

# Justera kolumnbredd automatiskt
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

# Spara resultatet
wb.save("rensad_output_formaterad.xlsx")

print("✔️ Klar! Filen 'rensad_output_formaterad.xlsx' har sparats.")
